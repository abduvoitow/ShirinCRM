from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from .forms import *
from django.http import JsonResponse, HttpResponse
from datetime import date
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from openpyxl.styles import Alignment, Font, Border, Side
from django.contrib import messages
from django.db.models import BooleanField, ExpressionWrapper, Q


# Login va logout funksiyalari
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Login yoki parol noto‘g‘ri, qaytadan urinib ko‘ring')
    return render(request, 'center/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')



# @login_required
# def home(request):# Bosh sahifa, Guruhlar ro'yxati
#     query = request.GET.get('q', '')
#     groups = Group.objects.filter(is_archived=False).order_by('-start_date')
#     today = date.today()
#     if query:
#         students = Student.objects.filter(
#             Q(full_name__icontains=query) |
#             Q(phone__icontains=query) |
#             Q(group__name__icontains=query)
#         ).select_related('group').prefetch_related('payments')
#     else:
#         students = Student.objects.none()

#     for group in groups:
#         group_students = group.students.all()
#         group.student_count = group_students.count()
#         group.paid_count = group_students.filter(payments__isnull=False).distinct().count()

#     return render(request, 'center/home.html', {
#         'groups': groups,
#         'students': students,
#         'query': query,
#         'today': today, 
#     })


@login_required
def home(request):
    query = request.GET.get('q', '')
    groups = Group.objects.filter(is_archived=False).order_by('-start_date')
    today = date.today()

    if query:
        students = Student.objects.filter(
            Q(full_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(group__name__icontains=query)
        ).select_related('group').prefetch_related('payments'
        ).annotate(
            is_group_archived=ExpressionWrapper(Q(group__is_archived=True), output_field=BooleanField())
        ).order_by('is_group_archived', 'full_name')
    else:
        students = Student.objects.none()

    for group in groups:
        group_students = group.students.all()
        group.student_count = group_students.count()
        group.paid_count = group_students.filter(payments__isnull=False).distinct().count()

    return render(request, 'center/home.html', {
        'groups': groups,
        'students': students,
        'query': query,
        'today': today,
    })



def create_group(request):  # Guruh qo'shish
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            messages.success(request, f"Yangi guruh {group.name} muvaffaqiyatli yaratildi!")
            return redirect('group_detail', group_id=group.id)
    else:
        form = GroupForm()
    return render(request, 'center/create_group.html', {'form': form})



def group_detail(request, group_id):#  Guruh sahifasi
    group = get_object_or_404(Group, pk=group_id)
    students = group.students.all()
    return render(request, 'center/group_detail.html', {'group': group, 'students': students})


def edit_group(request, group_id):# Guruhni tahrirlash
    group = get_object_or_404(Group, id=group_id)
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, f"{group.name} guruhi muvaffaqiyatli tahrirlandi!")
            return redirect('group_detail', group_id=group.id)
    else:
        form = GroupForm(instance=group)
    return render(request, 'center/edit_group.html', {'form': form, 'group': group})


def delete_group(request, group_id):# Guruhni o‘chirish
    group = get_object_or_404(Group, id=group_id)
    if request.method == 'POST':
        group.delete()
        messages.success(request, f"{group.name} guruhi o'chirildi!")
        return redirect('home')
    return render(request, 'center/delete_group.html', {'group': group})








def search_students(request):# Talabalarni qidirish
    query = request.GET.get('q', '')
    if query:
        students = Student.objects.filter(full_name__icontains=query).select_related('group')
        results = [{
            'full_name': student.full_name,
            'group_id': student.group.id,
        } for student in students]
    else:
        results = []
    return JsonResponse(results, safe=False)








def add_student(request, group_id):  # O'quvchi qo‘shish
    group = get_object_or_404(Group, id=group_id)
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.group = group
            student.save()
            messages.success(request, f"{group.name} guruhiga {student.full_name} qo‘shildi!")
            return redirect('group_detail', group_id=group.id)
    else:
        form = StudentForm()
    return render(request, 'center/add_student.html', {'form': form, 'group': group})


def edit_student(request, student_id):# O'quvchini tahrirlash
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f"{student.full_name} ma'lumotlari tahrirlandi!")
            return redirect('group_detail', group_id=student.group.id)
    else:
        form = StudentForm(instance=student)
    return render(request, 'center/edit_student.html', {'form': form, 'student': student})

def delete_student(request, student_id):# O'quvchini o‘chirish
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        group_id = student.group.id
        student.delete()
        messages.success(request, f"{student.full_name} o'chirildi!")
        return redirect('group_detail', group_id=group_id)
    return render(request, 'center/delete_student.html', {'student': student})





def add_payment(request, student_id):# To‘lov qo‘shish
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.student = student
            payment.save()
            messages.success(request, f"{student.full_name} to'lovi saqlandi!")
            return redirect('group_detail', group_id=student.group.id)
    else:
        form = PaymentForm()
    return render(request, 'center/add_payment.html', {'form': form, 'student': student})


def edit_payment(request, payment_id):# To‘lovni tahrirlash
    payment = get_object_or_404(Payment, id=payment_id)
    student = payment.student
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, f"{student.full_name} to'lovi tahrirlandi!")
            return redirect('group_detail', group_id=student.group.id)
    else:
        form = PaymentForm(instance=payment)
    return render(request, 'center/edit_payment.html', {'form': form, 'student': student, 'payment': payment})


def delete_payment(request, payment_id):# To‘lovni o‘chirish
    payment = get_object_or_404(Payment, id=payment_id)
    student = payment.student
    if request.method == 'POST':
        payment.delete()
        messages.success(request, f"{student.full_name} to'lovi o'chirildi!")
        return redirect('group_detail', group_id=student.group.id)
    return redirect('edit_payment', payment_id=payment.id)



# Davomatlarni ko‘rish
def group_attendance(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    students = group.students.all()
    return render(request, 'center/group_attendance.html', {'group': group, 'students': students})

def attendance_view(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    students = group.students.all()
    today = date.today()

    # POST so‘rovi orqali bugungi davomatni saqlash
    if request.method == 'POST':
        for student in students:
            present = request.POST.get(f'attendance_{student.id}') == 'on'
            Attendance.objects.update_or_create(
                student=student,
                date=today,
                defaults={'is_present': present}
            )
        messages.success(request, "Bugungi davomat muvaffaqiyatli saqlandi.")
        return redirect('attendance_view', group_id=group.id)

    # Barcha sanalarni olish
    all_dates_qs = Attendance.objects.filter(student__group=group).values_list('date', flat=True).distinct().order_by('-date')
    all_dates = list(all_dates_qs)

    # Agar bugungi sana mavjud bo‘lsa, all_dates dan olib tashlaymiz
    if today in all_dates:
        all_dates.remove(today)

    # Har bir student uchun davomat yozuvlarini olish
    attendance_data = {}
    for student in students:
        records = Attendance.objects.filter(student=student)
        attendance_data[student] = {record.date: record.is_present for record in records}

    return render(request, 'center/attendance.html', {
        'group': group,
        'students': students,
        'attendance_data': attendance_data,
        'all_dates': all_dates,
        'today': today
    })





#  Excel, yuklab olish
def export_group_excel(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    students = group.students.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{group.name} o'quvchilari"

    headers = ['№', 'F.I.SH', 'Telefon raqami', 'Seriya', 'Raqam', 'JSHSHIR', 'To‘lov', 'Shartnoma']

    # Borderlar
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # style
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f'{col_letter}1']
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # widthlar
        if header == 'F.I.SH':
            sheet.column_dimensions[col_letter].width = 30
        else:
            sheet.column_dimensions[col_letter].width = 15

    # th
    for row_num, (index, student) in enumerate(enumerate(students, 1), 2):
        data = [
            index,
            student.full_name,
            student.phone,
            student.passport_series,
            student.passport_number,
            student.jshr,
            "+" if student.payments.exists() else "-",
            "+" if student.has_contract else "-"
        ]
        for col_num, value in enumerate(data, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f'{col_letter}{row_num}']
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={group.name} malaka oshirish.xlsx'
    workbook.save(response)
    return response



# Shartnoma
def toggle_contract(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.has_contract = not student.has_contract
    student.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))



def toggle_archive(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    group.is_archived = not group.is_archived
    group.save()
    return redirect('home')


@login_required
def archived_groups(request):
    groups = Group.objects.filter(is_archived=True)
    for group in groups:
        group_students = group.students.all()
        group.student_count = group_students.count()
        group.paid_count = group_students.filter(payments__isnull=False).distinct().count()

    return render(request, 'center/archived_groups.html', {'groups': groups})



def unarchive_group(request, group_id):
    group = Group.objects.get(id=group_id)
    group.is_archived = False
    group.save()
    messages.success(request, f"{group.name} guruhi arxivdan chiqarildi.")
    return redirect('archived_groups') 


def archive_group(request, group_id):
    group = Group.objects.get(id=group_id)
    group.is_archived = True
    group.save()
    messages.success(request, f"{group.name} guruhi  arxivlandi.")
    return redirect('home')